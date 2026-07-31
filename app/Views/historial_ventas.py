from datetime import datetime
import os
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk

from app.models.venta import Venta
from app.printing.ticket_printer import TicketPrinter

# Intentar importar openpyxl para la exportación a Excel
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class HistorialVentasWindow(ctk.CTkToplevel):

    def __init__(self, parent):
        super().__init__(parent)

        self.title("Historial y Reporte de Ventas")
        self.geometry("1100x700")
        self.minsize(950, 600)

        # Hacer la ventana modal
        self.transient(parent)
        self.grab_set()

        # Estructura principal
        self._crear_encabezado()
        self._crear_filtros()
        self._crear_tabla()
        self._crear_resumen_y_acciones()

        # Cargar datos iniciales
        self.cargar_ventas_del_dia()

    # =====================================================
    # COMPONENTES DE LA INTERFAZ
    # =====================================================

    def _crear_encabezado(self):
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(15, 5))

        lbl_titulo = ctk.CTkLabel(
            header_frame,
            text="📋 Historial, Corte y Reportes de Ventas",
            font=ctk.CTkFont(size=22, weight="bold"),
        )
        lbl_titulo.pack(side="left")

    def _crear_filtros(self):
        filter_frame = ctk.CTkFrame(self)
        filter_frame.pack(fill="x", padx=20, pady=10)

        # Buscador por ID o Método
        lbl_buscar = ctk.CTkLabel(
            filter_frame, text="🔍 Buscar (Folio/Método):"
        )
        lbl_buscar.grid(row=0, column=0, padx=(15, 5), pady=12, sticky="w")

        self.txt_buscar = ctk.CTkEntry(
            filter_frame, placeholder_text="Ej. 105 o EFECTIVO...", width=180
        )
        self.txt_buscar.grid(row=0, column=1, padx=5, pady=12)
        self.txt_buscar.bind("<KeyRelease>", self._filtrar_por_busqueda)

        # Filtro Fecha Inicio
        lbl_inicio = ctk.CTkLabel(
            filter_frame, text="📅 Desde (YYYY-MM-DD):"
        )
        lbl_inicio.grid(row=0, column=2, padx=(15, 5), pady=12, sticky="w")

        self.txt_fecha_inicio = ctk.CTkEntry(
            filter_frame, placeholder_text="2026-01-01", width=110
        )
        self.txt_fecha_inicio.grid(row=0, column=3, padx=5, pady=12)

        # Filtro Fecha Fin
        lbl_fin = ctk.CTkLabel(filter_frame, text="Hasta:")
        lbl_fin.grid(row=0, column=4, padx=(10, 5), pady=12, sticky="w")

        self.txt_fecha_fin = ctk.CTkEntry(
            filter_frame, placeholder_text="2026-12-31", width=110
        )
        self.txt_fecha_fin.grid(row=0, column=5, padx=5, pady=12)

        # Botón Filtrar Rango
        btn_filtrar = ctk.CTkButton(
            filter_frame,
            text="Filtrar",
            width=80,
            command=self._filtrar_por_rango_fechas,
        )
        btn_filtrar.grid(row=0, column=6, padx=5, pady=12)

        # Botón Restablecer (Hoy)
        btn_hoy = ctk.CTkButton(
            filter_frame,
            text="Ver Hoy",
            width=80,
            fg_color="gray",
            hover_color="#555555",
            command=self.cargar_ventas_del_dia,
        )
        btn_hoy.grid(row=0, column=7, padx=(0, 15), pady=12)

    def _crear_tabla(self):
        table_frame = ctk.CTkFrame(self)
        table_frame.pack(fill="both", expand=True, padx=20, pady=5)

        # Estilos para el Treeview de Tkinter adaptados a CustomTkinter (Modo Oscuro)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background="#2b2b2b",
            foreground="white",
            fieldbackground="#2b2b2b",
            rowheight=28,
            font=("Arial", 10),
        )
        style.configure(
            "Treeview.Heading",
            background="#1f1f1f",
            foreground="white",
            font=("Arial", 11, "bold"),
        )
        style.map("Treeview", background=[("selected", "#1f538d")])

        columnas = ("id", "fecha", "metodo", "subtotal", "descuento", "total")

        self.tabla = ttk.Treeview(
            table_frame,
            columns=columnas,
            show="headings",
            selectmode="browse",
        )

        self.tabla.heading("id", text="Folio")
        self.tabla.heading("fecha", text="Fecha y Hora")
        self.tabla.heading("metodo", text="Método de Pago")
        self.tabla.heading("subtotal", text="Subtotal")
        self.tabla.heading("descuento", text="Descuento")
        self.tabla.heading("total", text="Total")

        self.tabla.column("id", width=70, anchor="center")
        self.tabla.column("fecha", width=180, anchor="center")
        self.tabla.column("metodo", width=130, anchor="center")
        self.tabla.column("subtotal", width=110, anchor="e")
        self.tabla.column("descuento", width=110, anchor="e")
        self.tabla.column("total", width=120, anchor="e")

        scrollbar = ttk.Scrollbar(
            table_frame, orient="vertical", command=self.tabla.yview
        )
        self.tabla.configure(yscroll=scrollbar.set)

        self.tabla.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Doble clic para ver detalle
        self.tabla.bind("<Double-1>", lambda event: self.ver_detalle_venta())

    def _crear_resumen_y_acciones(self):
        bottom_frame = ctk.CTkFrame(self)
        bottom_frame.pack(fill="x", padx=20, pady=(5, 15))

        # Panel de totales acumulados
        self.lbl_conteo = ctk.CTkLabel(
            bottom_frame,
            text="Ventas: 0",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        self.lbl_conteo.pack(side="left", padx=15, pady=15)

        self.lbl_total_periodo = ctk.CTkLabel(
            bottom_frame,
            text="Total Período: $0.00",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#2FA572",
        )
        self.lbl_total_periodo.pack(side="left", padx=15, pady=15)

        # Botones de Acción Profesional
        btn_excel = ctk.CTkButton(
            bottom_frame,
            text="📦 Exportar Excel",
            fg_color="#1E7145",
            hover_color="#134B2E",
            command=self.exportar_excel,
        )
        btn_excel.pack(side="right", padx=5, pady=15)

        btn_corte = ctk.CTkButton(
            bottom_frame,
            text="💵 Corte de Caja",
            fg_color="#D97706",
            hover_color="#B45309",
            command=self.abrir_corte_caja,
        )
        btn_corte.pack(side="right", padx=5, pady=15)

        btn_reimprimir = ctk.CTkButton(
            bottom_frame,
            text="🖨️ Reimprimir",
            fg_color="#1f538d",
            hover_color="#14375e",
            command=self.reimprimir_ticket,
        )
        btn_reimprimir.pack(side="right", padx=5, pady=15)

        btn_detalle = ctk.CTkButton(
            bottom_frame,
            text="👁️ Ver Detalle",
            fg_color="#3B8ED0",
            hover_color="#36719F",
            command=self.ver_detalle_venta,
        )
        btn_detalle.pack(side="right", padx=5, pady=15)

    # =====================================================
    # LÓGICA Y CONSULTAS DE DATOS
    # =====================================================

    def _poblar_tabla(self, registros):
        """Limpia la tabla actual y la llena con la lista de registros enviada."""
        for row in self.tabla.get_children():
            self.tabla.delete(row)

        suma_total = 0.0

        for venta in registros:
            v_dict = dict(venta) if hasattr(venta, "keys") else venta
            subtotal = f"${v_dict['subtotal']:.2f}"
            descuento = f"${v_dict['descuento']:.2f}"
            total = f"${v_dict['total']:.2f}"
            suma_total += float(v_dict["total"])

            self.tabla.insert(
                "",
                "end",
                values=(
                    v_dict["id"],
                    v_dict["fecha"],
                    v_dict["metodo_pago"],
                    subtotal,
                    descuento,
                    total,
                ),
            )

        self.lbl_conteo.configure(text=f"Ventas en lista: {len(registros)}")
        self.lbl_total_periodo.configure(
            text=f"Total Período: ${suma_total:.2f}"
        )

    def cargar_ventas_del_dia(self):
        """Carga las ventas realizadas el día de hoy por defecto."""
        self.txt_buscar.delete(0, "end")
        self.txt_fecha_inicio.delete(0, "end")
        self.txt_fecha_fin.delete(0, "end")

        hoy = datetime.now().strftime("%Y-%m-%d")
        self.txt_fecha_inicio.insert(0, hoy)
        self.txt_fecha_fin.insert(0, hoy)

        ventas = Venta.obtener_ventas_del_dia()
        self._poblar_tabla(ventas)

    def _filtrar_por_busqueda(self, event=None):
        texto = self.txt_buscar.get().strip()
        if not texto:
            self.cargar_ventas_del_dia()
            return

        ventas = Venta.buscar(texto)
        self._poblar_tabla(ventas)

    def _filtrar_por_rango_fechas(self):
        f_inicio = self.txt_fecha_inicio.get().strip()
        f_fin = self.txt_fecha_fin.get().strip()

        if not f_inicio or not f_fin:
            messagebox.showwarning(
                "Atención",
                "Por favor ingresa ambas fechas en formato YYYY-MM-DD.",
            )
            return

        try:
            ventas = Venta.obtener_por_rango(f_inicio, f_fin)
            self._poblar_tabla(ventas)
        except Exception as e:
            messagebox.showerror(
                "Error", f"Ocurrió un error al filtrar por fechas: {str(e)}"
            )

    def _obtener_id_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            messagebox.showwarning(
                "Atención", "Por favor selecciona una venta de la lista."
            )
            return None
        valores = self.tabla.item(seleccion[0], "values")
        return valores[0]  # Retorna el folio (id)

    # =====================================================
    # DETALLE Y REIMPRESIÓN
    # =====================================================

    def ver_detalle_venta(self):
        venta_id = self._obtener_id_seleccionado()
        if not venta_id:
            return

        venta, detalles = Venta.obtener_ticket(venta_id)
        if not venta:
            messagebox.showerror(
                "Error", "No se encontró el detalle de la venta especificada."
            )
            return

        venta = dict(venta) if hasattr(venta, "keys") else venta

        # Ventana Modal de Detalle con CustomTkinter
        dialogo = ctk.CTkToplevel(self)
        dialogo.title(f"Detalle de Venta #{venta_id}")
        dialogo.geometry("580x480")
        dialogo.transient(self)
        dialogo.grab_set()

        # Cabecera
        lbl_info = ctk.CTkLabel(
            dialogo,
            text=(
                f"Folio: #{venta['id']}  |  Fecha: {venta['fecha']}\n"
                f"Método de Pago: {venta['metodo_pago']}"
            ),
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        lbl_info.pack(pady=10)

        # Tabla de productos
        table_frame = ctk.CTkFrame(dialogo)
        table_frame.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("prod", "cant", "precio", "subtotal")
        tabla_det = ttk.Treeview(table_frame, columns=cols, show="headings")
        tabla_det.heading("prod", text="Producto")
        tabla_det.heading("cant", text="Cant.")
        tabla_det.heading("precio", text="P. Unit")
        tabla_det.heading("subtotal", text="Subtotal")

        tabla_det.column("prod", width=220)
        tabla_det.column("cant", width=70, anchor="center")
        tabla_det.column("precio", width=80, anchor="e")
        tabla_det.column("subtotal", width=90, anchor="e")

        for item in detalles:
            item_dict = dict(item) if hasattr(item, "keys") else item
            unidad = item_dict.get("unidad", "")
            texto_cantidad = f"{item_dict['cantidad']} {unidad}".strip()

            tabla_det.insert(
                "",
                "end",
                values=(
                    item_dict["producto_nombre"],
                    texto_cantidad,
                    f"${item_dict['precio_unitario']:.2f}",
                    f"${item_dict['subtotal']:.2f}",
                ),
            )

        tabla_det.pack(fill="both", expand=True)

        # Totales
        lbl_totales = ctk.CTkLabel(
            dialogo,
            text=(
                f"Subtotal: ${venta['subtotal']:.2f}  |  "
                f"Descuento: ${venta['descuento']:.2f}  |  "
                f"Total: ${venta['total']:.2f}"
            ),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#2FA572",
        )
        lbl_totales.pack(pady=10)

    def reimprimir_ticket(self):
        venta_id = self._obtener_id_seleccionado()
        if not venta_id:
            return

        venta, detalles = Venta.obtener_ticket(venta_id)
        if not venta:
            messagebox.showerror(
                "Error", "No se pudo recuperar la información de la venta."
            )
            return

        venta = dict(venta) if hasattr(venta, "keys") else venta

        try:
            # Adaptar los detalles al formato que espera TicketPrinter
            carrito = []
            for item in detalles:
                item_dict = dict(item) if hasattr(item, "keys") else item
                carrito.append({
                    "nombre": item_dict["producto_nombre"],
                    "cantidad": item_dict["cantidad"],
                    "precio": item_dict["precio_unitario"],
                    "subtotal": item_dict["subtotal"],
                })

            # Invocación a tu módulo TicketPrinter
            TicketPrinter.imprimir_ticket(
                venta["id"],
                carrito,
                venta["total"],
                venta.get("monto_recibido", venta["total"]),
                venta.get("cambio", 0.0),
                venta["metodo_pago"],
            )

            messagebox.showinfo(
                "Reimpresión",
                f"Ticket #{venta_id} enviado a la impresora con éxito.",
            )

        except Exception as e:
            messagebox.showerror(
                "Error de Impresión",
                f"No se pudo enviar la reimpresión: {str(e)}",
            )

    # =====================================================
    # NUEVAS FUNCIONALIDADES PROFESIONALES
    # =====================================================

    def exportar_excel(self):
        """Exporta las ventas visibles y sus detalles a un archivo Excel profesional."""
        if not EXCEL_AVAILABLE:
            messagebox.showerror(
                "Librería Faltante",
                "Se requiere 'openpyxl' para exportar a Excel.\nInstala ejecutando: pip install openpyxl",
            )
            return

        f_ini = self.txt_fecha_inicio.get().strip()
        f_fin = self.txt_fecha_fin.get().strip()

        try:
            ventas = Venta.obtener_por_rango(f_ini, f_fin)
        except Exception:
            ventas = Venta.obtener_ventas_del_dia()

        if not ventas:
            messagebox.showwarning(
                "Sin Datos", "No hay ventas cargadas para exportar."
            )
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Guardar Historial de Ventas",
            initialfile=f"Reporte_Ventas_{f_ini}_al_{f_fin}.xlsx",
        )

        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()

            # Hoja 1: Resumen de Ventas
            ws1 = wb.active
            ws1.title = "Ventas Generales"

            headers_v = [
                "Folio",
                "Fecha/Hora",
                "Método Pago",
                "Subtotal",
                "Descuento",
                "Total",
                "Recibido",
                "Cambio",
            ]
            ws1.append(headers_v)

            header_fill = PatternFill(
                start_color="1F4E79", end_color="1F4E79", fill_type="solid"
            )
            header_font = Font(
                name="Calibri", size=11, bold=True, color="FFFFFF"
            )
            border_thin = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            for col_num in range(1, len(headers_v) + 1):
                cell = ws1.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            for v in ventas:
                v_dict = dict(v) if hasattr(v, "keys") else v
                ws1.append([
                    v_dict["id"],
                    v_dict["fecha"],
                    v_dict["metodo_pago"],
                    v_dict["subtotal"],
                    v_dict["descuento"],
                    v_dict["total"],
                    v_dict.get("monto_recibido", 0.0),
                    v_dict.get("cambio", 0.0),
                ])

            # Formatear números en moneda
            for row in range(2, ws1.max_row + 1):
                for col_idx in [4, 5, 6, 7, 8]:
                    cell = ws1.cell(row=row, column=col_idx)
                    cell.number_format = '"$"#,##0.00'
                    cell.border = border_thin

            # Hoja 2: Detalle de Artículos
            ws2 = wb.create_sheet(title="Detalle Artículos")
            headers_d = [
                "Folio Venta",
                "Producto",
                "Cantidad",
                "Precio Unitario",
                "Subtotal",
            ]
            ws2.append(headers_d)

            for col_num in range(1, len(headers_d) + 1):
                cell = ws2.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            for v in ventas:
                v_dict = dict(v) if hasattr(v, "keys") else v
                _, detalles = Venta.obtener_ticket(v_dict["id"])
                for d in detalles:
                    d_dict = dict(d) if hasattr(d, "keys") else d
                    ws2.append([
                        d_dict["venta_id"],
                        d_dict["producto_nombre"],
                        d_dict["cantidad"],
                        d_dict["precio_unitario"],
                        d_dict["subtotal"],
                    ])

            for row in range(2, ws2.max_row + 1):
                ws2.cell(row=row, column=4).number_format = '"$"#,##0.00'
                ws2.cell(row=row, column=5).number_format = '"$"#,##0.00'

            # Autoajuste de columnas
            for ws in [ws1, ws2]:
                for col in ws.columns:
                    max_len = max(
                        len(str(cell.value or "")) for cell in col
                    )
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(
                        max_len + 3, 12
                    )

            wb.save(filepath)
            messagebox.showinfo(
                "Éxito", f"Reporte guardado correctamente en:\n{filepath}"
            )

        except Exception as e:
            messagebox.showerror(
                "Error de Exportación", f"No se pudo generar el Excel: {str(e)}"
            )

    def abrir_corte_caja(self):
        """Abre la ventana de arqueo y corte de caja del día de hoy."""
        hoy = datetime.now().strftime("%Y-%m-%d")
        try:
            ventas_hoy = Venta.obtener_por_rango(hoy, hoy)
        except Exception:
            ventas_hoy = Venta.obtener_ventas_del_dia()

        ventas_dict = [dict(v) if hasattr(v, "keys") else v for v in ventas_hoy]

        total_efectivo = sum(
            float(v["total"])
            for v in ventas_dict
            if str(v["metodo_pago"]).upper() == "EFECTIVO"
        )
        total_otros = sum(
            float(v["total"])
            for v in ventas_dict
            if str(v["metodo_pago"]).upper() != "EFECTIVO"
        )
        total_general = total_efectivo + total_otros

        top = ctk.CTkToplevel(self)
        top.title("💵 Corte de Caja - Arqueo")
        top.geometry("420x460")
        top.transient(self)
        top.grab_set()

        lbl_tit = ctk.CTkLabel(
            top,
            text="CORTE DE CAJA DEL DÍA",
            font=ctk.CTkFont(size=18, weight="bold"),
        )
        lbl_tit.pack(pady=(15, 10))

        # Cuadro Informativo
        info_frame = ctk.CTkFrame(top)
        info_frame.pack(fill="x", padx=20, pady=5)

        ctk.CTkLabel(
            info_frame, text=f"Ventas en Efectivo: ${total_efectivo:.2f}"
        ).pack(anchor="w", padx=15, pady=3)
        ctk.CTkLabel(
            info_frame, text=f"Ventas Tarjeta/Otros: ${total_otros:.2f}"
        ).pack(anchor="w", padx=15, pady=3)
        ctk.CTkLabel(
            info_frame,
            text=f"Total Vendido Hoy: ${total_general:.2f}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#2FA572",
        ).pack(anchor="w", padx=15, pady=5)

        # Entradas
        input_frame = ctk.CTkFrame(top)
        input_frame.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(input_frame, text="Fondo Inicial en Caja ($):").pack(
            anchor="w", padx=15, pady=(8, 2)
        )
        txt_fondo = ctk.CTkEntry(input_frame, placeholder_text="0.00")
        txt_fondo.pack(fill="x", padx=15, pady=(0, 8))
        txt_fondo.insert(0, "0.00")

        ctk.CTkLabel(input_frame, text="Efectivo Físico Contado ($):").pack(
            anchor="w", padx=15, pady=(8, 2)
        )
        txt_efectivo_real = ctk.CTkEntry(input_frame, placeholder_text="0.00")
        txt_efectivo_real.pack(fill="x", padx=15, pady=(0, 8))

        lbl_resultado = ctk.CTkLabel(
            top, text="", font=ctk.CTkFont(size=13, weight="bold")
        )
        lbl_resultado.pack(pady=10)

        def calcular_diferencia():
            try:
                fondo = float(txt_fondo.get().strip() or 0)
                efectivo_real = float(txt_efectivo_real.get().strip() or 0)
                esperado = total_efectivo + fondo
                diferencia = efectivo_real - esperado

                if diferencia == 0:
                    lbl_resultado.configure(
                        text="✅ CAJA CUADRADA PERFECTAMENTE", text_color="#2FA572"
                    )
                elif diferencia > 0:
                    lbl_resultado.configure(
                        text=f"🔵 SOBRANTE: ${diferencia:.2f}",
                        text_color="#3B8ED0",
                    )
                else:
                    lbl_resultado.configure(
                        text=f"🔴 FALTANTE: ${abs(diferencia):.2f}",
                        text_color="#E53E3E",
                    )
            except ValueError:
                messagebox.showerror(
                    "Error", "Ingrese valores numéricos válidos."
                )

        btn_calc = ctk.CTkButton(
            top,
            text="Calcular Diferencia",
            command=calcular_diferencia,
            fg_color="#D97706",
            hover_color="#B45309",
        )
        btn_calc.pack(pady=5)